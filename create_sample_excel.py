#!/usr/bin/env python3
"""
Script to create a sample Excel file for testing the job extraction application.
"""
import os
from openpyxl import Workbook

def create_sample_excel():
    # Ensure data directory exists
    os.makedirs('data', exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Companies'

    # Add headers
    ws['A1'] = 'id'
    ws['B1'] = 'location'
    ws['C1'] = 'website'
    ws['D1'] = 'websiteToJobs'

    # Add sample data
    ws['A2'] = 1
    ws['B2'] = 'Berlin, Germany'
    ws['C2'] = 'https://www.example-company.com'
    ws['D2'] = 'https://www.example-company.com/careers'

    ws['A3'] = 2
    ws['B3'] = 'Munich, Germany'
    ws['C3'] = 'https://www.another-company.com'
    ws['D3'] = 'https://www.another-company.com/jobs'

    ws['A4'] = 3
    ws['B4'] = 'Hamburg, Germany'
    ws['C4'] = 'https://www.tech-startup.com'
    ws['D4'] = 'https://www.tech-startup.com/careers'

    wb.save('data/Landratsamt.xlsx')
    print('Sample Excel file created successfully at data/Landratsamt.xlsx')

if __name__ == '__main__':
    create_sample_excel()
